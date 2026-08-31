import openpyxl, json, re
wb = openpyxl.load_workbook('BOT.xlsx')
out={}

# Sectors from MSP2-03 (rows 14..77 step 3)
ws=wb['MSP2-03']
sectors=[]
for r in range(14,78):
    v=ws.cell(r,2).value
    if v and re.match(r'^\d+\.',str(v).strip()):
        sectors.append(re.sub(r'^\d+\.\s*','',str(v).strip()))
out['sectors']=sectors

# Loan types MSP2-04 rows 14..27
ws=wb['MSP2-04']
lt=[]
for r in range(14,28):
    v=ws.cell(r,2).value
    if v: lt.append({'row':r,'label':str(v).strip()})
out['loan_types']=lt

# Provision matrices MSP2-03
ws=wb['MSP2-03']
out['provisioning_standard']=[{'band':str(ws.cell(r,2).value).strip(),'rate':float(ws.cell(r,3).value)} for r in range(89,94)]
out['provisioning_housing']=[{'band':str(ws.cell(r,2).value).strip(),'rate':float(ws.cell(r,3).value)} for r in range(96,99)]
out['classification_columns']=[str(wb['MSP2-03'].cell(12,c).value or wb['MSP2-03'].cell(11,c).value).strip() for c in range(5,10)]

# Complaint natures MSP2-06 row12 E..J ; particulars col B 14..22
ws=wb['MSP2-06']
out['complaint_natures']=[str(ws.cell(12,c).value).strip() for c in range(5,11)]
out['complaint_rows']=[re.sub(r'\s+',' ',str(ws.cell(r,2).value).strip()) for r in range(14,23)]

# Banks
ws=wb['Static Information']
out['banks_list_A']=[str(ws.cell(r,2).value).strip() for r in range(3,59) if ws.cell(r,2).value]
out['banks_list_B_agent']=[str(ws.cell(r,7).value).strip() for r in range(3,55) if ws.cell(r,7).value]

# MNOs MSP2-07 61..66
ws=wb['MSP2-07']
out['mnos']=[str(ws.cell(r,2).value).strip() for r in range(61,67) if ws.cell(r,2).value]
out['msp2_07_sections']={'banks_tz':'rows 15-42','msps':'rows 45-58','mnos':'rows 61-66','banks_abroad':'rows 72-77'}

# MSP2-10 geography
ws=wb['MSP2-10']
geo=[];cur=None;area=None
for r in range(14,243):
    label=ws.cell(r,2).value
    if not label: continue
    label=str(label).strip()
    c=ws.cell(r,3).value
    if label.startswith('A.') or label.startswith('B.'):
        area=label; continue
    if label.startswith('Total') or label.startswith('Grand'): continue
    if isinstance(c,str) and c.startswith('=SUM'):
        cur={'area':area,'region':label,'districts':[]}; geo.append(cur)
    elif cur is not None:
        cur['districts'].append(label)
out['geography']=geo
out['msp2_10_dimensions']={'age_bands':['Up to 35 Years','Above 35 Years'],'genders':['Female','Male'],
 'measures':['Number of Branches','Number of Employees','Compulsory savings','Number of Borrowers','Number of Loans','Outstanding Loans']}

json.dump(out,open('bot_reference.json','w'),indent=1)
print('sectors',len(sectors)); print(sectors)
print('\nloan_types',len(lt)); [print(' ',x['row'],x['label']) for x in lt]
print('\nstd prov',out['provisioning_standard'])
print('housing prov',out['provisioning_housing'])
print('class cols',out['classification_columns'])
print('\ncomplaint natures',out['complaint_natures'])
print('complaint rows:'); [print('  -',x) for x in out['complaint_rows']]
print('\nmnos',out['mnos'])
print('banksA',len(out['banks_list_A']),'banksB',len(out['banks_list_B_agent']))
print('\ngeography regions:',len(geo),'districts:',sum(len(g['districts']) for g in geo))
for g in geo[:4]: print('  ',g['area'],'|',g['region'],'->',len(g['districts']),g['districts'][:4])
print('  areas:',sorted({g['area'] for g in geo}))
