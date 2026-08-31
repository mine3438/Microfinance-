"""Derive the cell map for BOT's MSP2 workbook, and emit it as a migration.

Where a figure goes in BOT's own spreadsheet is a fact about that spreadsheet,
not arithmetic this system is entitled to invent. So it is read off the template
and seeded as data: if BOT inserts a line in a future revision, the fix is a
migration re-run of this script, not a patch to the exporter.

Run from `docs/reference`; the template lives with the code that reads it,
in `apps/api/assets`:

    python3 extract-bot-cells.py > ../../db/migrations/0017_form_cells.sql

The taxonomy codes are parsed out of migration 0004 rather than re-slugified
here, so a name this script cannot match to a seeded row is a hard failure at
generation time instead of a silently missing line in a filed return.
"""

import re
import sys
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
TEMPLATE = HERE / ".." / ".." / "apps" / "api" / "assets" / "BOT-MSP2-template.xlsx"
REFERENCE_MIGRATION = HERE / ".." / ".." / "db" / "migrations" / "0004_bot_reference_data.sql"

# The first data row of every sheet is 14, and BOT's Sno column starts at 1
# there, so a line's row is its Sno plus this. Verified against all ten sheets.
SNO_ROW_OFFSET = 13

# Where each sheet carries the institution's name (row 2), its MSP code (row 4)
# and the quarter-end date (row 6). BOT's own template fills nine of these ten
# by formula from an external workbook that is not distributed with it, so the
# exporter writes every one of them literally.
IDENTITY_COLUMN = {
    "MSP2-01": "C",
    "MSP2-02": "C",
    "MSP2-03": "D",
    "MSP2-04": "D",
    "MSP2-05": "D",
    "MSP2-07": "F",
    "MSP2-08": "E",
    "MSP2-09": "D",
    "MSP2-10": "H",
}

SHEET_NAME = {code: ("MSP2_01" if code == "MSP2-01" else code) for code in IDENTITY_COLUMN}

# Column keys are the names this system's own wire shapes use, so the exporter
# reads as a translation between two vocabularies rather than as a lookup table
# of letters. Columns BOT computes are absent: the exporter has nothing to say
# about a cell that carries a formula.
COLUMNS = {
    "MSP2-01": {"amount": "C"},
    "MSP2-02": {"quarter_amount": "C", "year_to_date_amount": "D"},
    "MSP2-03": {
        "borrower_count": "C",
        "current": "E",
        "esm": "F",
        "substandard": "G",
        "doubtful": "H",
        "loss": "I",
        "written_off_during_period": "J",
    },
    "MSP2-04": {
        "borrower_count": "C",
        "total_outstanding": "D",
        "straight_line_lowest": "F",
        "straight_line_highest": "G",
        "reducing_balance_lowest": "I",
        "reducing_balance_highest": "J",
    },
    "MSP2-05": {"amount": "C"},
    "MSP2-07": {
        "counterparty": "B",
        "deposit_tzs": "C",
        "deposit_foreign_tzs_equivalent": "D",
        "borrowing_tzs": "F",
        "borrowing_foreign_tzs_equivalent": "G",
    },
    "MSP2-08": {"counterparty": "B", "balance": "C"},
    "MSP2-09": {
        "female_count": "C",
        "female_amount": "D",
        "male_count": "E",
        "male_amount": "F",
    },
    "MSP2-10": {
        "branch_count": "C",
        "employee_count": "D",
        "compulsory_savings": "E",
        "borrowers_up_to_35_female": "F",
        "borrowers_up_to_35_male": "G",
        "borrowers_above_35_female": "H",
        "borrowers_above_35_male": "I",
        "loans_up_to_35_female": "J",
        "loans_up_to_35_male": "K",
        "loans_above_35_female": "L",
        "loans_above_35_male": "M",
        "outstanding_up_to_35_female": "N",
        "outstanding_up_to_35_male": "O",
        "outstanding_above_35_female": "P",
        "outstanding_above_35_male": "Q",
    },
}

# MSP2-07 and MSP2-08 do not name their counterparties in advance: the rows are
# blank and the name is chosen from a dropdown beside the amount. So they are
# addressed as ranges to be filled in order, and the exporter refuses to file
# rather than silently drop a holding when a section runs out of rows.
SECTIONS = {
    "MSP2-07": [
        ("bank_tanzania", 15, 42),
        ("microfinance_service_provider", 45, 58),
        ("mno", 61, 68),
        ("bank_abroad", 72, 77),
    ],
    "MSP2-08": [("agent_banking", 15, 42)],
}


def fail(message: str) -> None:
    print(message, file=sys.stderr)
    raise SystemExit(1)


def seeded(table: str, columns: tuple[str, ...]) -> list[dict[str, str]]:
    """Read one INSERT ... VALUES block out of migration 0004."""
    sql = REFERENCE_MIGRATION.read_text(encoding="utf-8")
    start = sql.index(f"INSERT INTO reference.{table}")
    block = sql[start : sql.index(";", start)]
    rows: list[dict[str, str]] = []
    for line in block.splitlines():
        line = line.strip().rstrip(",")
        if not line.startswith("("):
            continue
        # Values are single-quoted literals, NULLs and integers; the doubled
        # quote is SQL's own escape and appears in no name here.
        parts = re.findall(r"'((?:[^']|'')*)'|\bNULL\b|(-?\d+)", line)
        flat = [(a or b).replace("''", "'") for a, b in parts]
        if len(flat) < len(columns):
            fail(f"{table}: cannot parse {line}")
        rows.append(dict(zip(columns, flat)))
    if not rows:
        fail(f"{table}: no seeded rows found")
    return rows


def label_of(sheet, row: int) -> str:
    value = sheet.cell(row, 2).value
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def numbered_label(sheet, row: int) -> str:
    """A sector label with BOT's leading ordinal removed."""
    return re.sub(r"^\d+\.\s*", "", label_of(sheet, row))


def literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def main() -> None:
    workbook = openpyxl.load_workbook(TEMPLATE)

    sectors = seeded("sectors", ("code", "sno", "name"))
    loan_types = seeded("loan_types", ("code", "sno", "name"))
    districts = seeded("districts", ("code", "region_code", "name", "council_type", "sort_order"))
    regions = seeded("regions", ("code", "area", "name", "sort_order"))

    by_sector_name = {row["name"]: row["code"] for row in sectors}
    by_loan_type_name = {row["name"]: row["code"] for row in loan_types}
    region_names = {row["code"]: row["name"] for row in regions}
    by_district = {(row["region_code"], row["name"]): row["code"] for row in districts}

    lines: list[str] = []

    # -- Sno-addressed forms -------------------------------------------------
    # MSP2-01, MSP2-02 and MSP2-05 are lists of numbered lines, and their Snos
    # are already seeded in `reference.form_lines`. Only the entered lines get a
    # row here; a computed line has nowhere to write by construction.
    sno_rows: list[tuple[str, int, int]] = []
    for form in ("MSP2-01", "MSP2-02", "MSP2-05"):
        sheet = workbook[SHEET_NAME[form]]
        column = COLUMNS[form]["amount" if form != "MSP2-02" else "quarter_amount"]
        for row in range(14, sheet.max_row + 1):
            sno = sheet.cell(row, 1).value
            if not isinstance(sno, int):
                continue
            value = sheet[f"{column}{row}"].value
            # A formula is BOT's own answer and this system has nothing to say
            # about it — unless it reads the workbook `[1]`, which is a file on
            # somebody's machine that BOT does not distribute with the template.
            # Those cells compute nothing here, so the exporter fills them.
            if isinstance(value, str) and value.startswith("=") and "[1]" not in value:
                continue
            if row != sno + SNO_ROW_OFFSET:
                fail(f"{form} Sno{sno} sits on row {row}, not {sno + SNO_ROW_OFFSET}")
            sno_rows.append((form, row, sno))

    # -- Taxonomy-addressed forms --------------------------------------------
    sector_rows: list[tuple[str, int, str]] = []
    for form in ("MSP2-03", "MSP2-09"):
        sheet = workbook[SHEET_NAME[form]]
        for row in range(14, sheet.max_row + 1):
            name = numbered_label(sheet, row)
            code = by_sector_name.get(name)
            if code is not None:
                sector_rows.append((form, row, code))
        found = len({code for f, _, code in sector_rows if f == form})
        if found != len(sectors):
            fail(f"{form}: matched {found} of {len(sectors)} sectors")

    loan_type_rows: list[tuple[str, int, str]] = []
    sheet = workbook["MSP2-04"]
    for row in range(14, sheet.max_row + 1):
        # BOT indents the two salaried sub-types with a bracketed letter.
        name = re.sub(r"^\([a-z]\)\s*", "", label_of(sheet, row))
        code = by_loan_type_name.get(name)
        if code is not None:
            loan_type_rows.append(("MSP2-04", row, code))
    if len(loan_type_rows) != len(loan_types):
        fail(f"MSP2-04: matched {len(loan_type_rows)} of {len(loan_types)} loan types")

    # -- MSP2-10: the districts BOT accepts entry for -------------------------
    # A region's own row carries a formula that sums its districts, so it is not
    # a cell this system writes. Walking the sheet keeps the region in hand, and
    # the district name alone is not unique across Tanzania.
    district_rows: list[tuple[str, int, str]] = []
    sheet = workbook["MSP2-10"]
    current_region: str | None = None
    region_by_name = {name: code for code, name in region_names.items()}
    for row in range(14, sheet.max_row + 1):
        name = label_of(sheet, row)
        if not name:
            continue
        first = sheet.cell(row, 3).value
        if isinstance(first, str) and first.startswith("="):
            current_region = region_by_name.get(name)
            continue
        if current_region is None:
            continue
        code = by_district.get((current_region, name))
        if code is not None:
            district_rows.append(("MSP2-10", row, code))
    if len(district_rows) != len(districts):
        fail(f"MSP2-10: matched {len(district_rows)} of {len(districts)} districts")

    # -- Emit -----------------------------------------------------------------
    lines.append("INSERT INTO reference.form_sheets")
    lines.append("  (form_code, sheet_name, sno_row_offset, identity_column) VALUES")
    body = [
        f"  ({literal(form)}, {literal(SHEET_NAME[form])}, {SNO_ROW_OFFSET}, "
        f"{literal(IDENTITY_COLUMN[form])})"
        for form in sorted(IDENTITY_COLUMN)
    ]
    lines.append(",\n".join(body) + ";")
    lines.append("")

    lines.append("INSERT INTO reference.form_columns (form_code, column_key, column_letter) VALUES")
    body = [
        f"  ({literal(form)}, {literal(key)}, {literal(letter)})"
        for form in sorted(COLUMNS)
        for key, letter in COLUMNS[form].items()
    ]
    lines.append(",\n".join(body) + ";")
    lines.append("")

    lines.append("INSERT INTO reference.form_sections")
    lines.append("  (form_code, section_key, first_row, last_row) VALUES")
    body = [
        f"  ({literal(form)}, {literal(key)}, {first}, {last})"
        for form in sorted(SECTIONS)
        for key, first, last in SECTIONS[form]
    ]
    lines.append(",\n".join(body) + ";")
    lines.append("")

    lines.append("INSERT INTO reference.form_rows (form_code, row_number, sno) VALUES")
    body = [f"  ({literal(form)}, {row}, {sno})" for form, row, sno in sno_rows]
    lines.append(",\n".join(body) + ";")
    lines.append("")

    lines.append("INSERT INTO reference.form_rows (form_code, row_number, sector_code) VALUES")
    body = [f"  ({literal(form)}, {row}, {literal(code)})" for form, row, code in sector_rows]
    lines.append(",\n".join(body) + ";")
    lines.append("")

    lines.append("INSERT INTO reference.form_rows (form_code, row_number, loan_type_code) VALUES")
    body = [f"  ({literal(form)}, {row}, {literal(code)})" for form, row, code in loan_type_rows]
    lines.append(",\n".join(body) + ";")
    lines.append("")

    lines.append("INSERT INTO reference.form_rows (form_code, row_number, district_code) VALUES")
    body = [f"  ({literal(form)}, {row}, {literal(code)})" for form, row, code in district_rows]
    lines.append(",\n".join(body) + ";")

    print("\n".join(lines))
    print(
        f"-- {len(sno_rows)} numbered lines, {len(sector_rows)} sector rows, "
        f"{len(loan_type_rows)} loan-type rows, {len(district_rows)} district rows.",
    )


if __name__ == "__main__":
    main()
